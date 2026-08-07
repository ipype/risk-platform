"""The same document as a workbook.

One sheet per section, and every figure written as a number with an Excel format rather
than as a pre-formatted string. A client who receives this will re-sum a column — that is
what a workbook is for — and a column of text that looks like currency makes them re-type
it first.

The style constants are deliberately duplicated from the register export rather than
imported from it: that module is a route, and a service importing a route is the wrong way
round. Six fills is a cheaper price than that dependency.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.report.model import (
    Callout,
    Document,
    KeyValues,
    MatrixBlock,
    Paragraph,
    Section,
    Table,
    excel_number_format,
    format_value,
)

__all__ = ["render_xlsx"]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
CAPTION_FONT = Font(bold=True, size=11, color="374151")
SUBTLE_FONT = Font(color="6B7280", size=10)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

TONE_FILL = {
    "warning": PatternFill("solid", fgColor="FEF3C7"),
    "method": PatternFill("solid", fgColor="DBEAFE"),
    "info": PatternFill("solid", fgColor="F3F4F6"),
}
TONE_LABEL = {"info": "NOTE", "warning": "WARNING", "method": "METHOD"}


def _hex(color: str | None) -> str | None:
    if not color:
        return None
    cleaned = str(color).lstrip("#")
    return cleaned if len(cleaned) in (6, 8) else None


def _sheet_title(document: Document, section: Section, used: set[str]) -> str:
    # Excel: 31 characters, and none of : \ / ? * [ ]
    base = "".join(c for c in section.title if c not in ':\\/?*[]')[:31].strip() or section.id
    title = base
    suffix = 2
    while title.lower() in used:
        title = f"{base[:28]} {suffix}"
        suffix += 1
    used.add(title.lower())
    return title


def _widths(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        letter = get_column_letter(i)
        current = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(current, min(width, 70))


def _write_table(ws: Worksheet, block: Table, row: int, currency: str) -> int:
    if block.caption:
        ws.cell(row=row, column=1, value=block.caption).font = CAPTION_FONT
        row += 1
    if not block.rows:
        ws.cell(row=row, column=1, value=block.empty_text).font = SUBTLE_FONT
        return row + 2

    for col, column in enumerate(block.columns, start=1):
        cell = ws.cell(row=row, column=col, value=column.label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    header_row = row
    row += 1

    for data_row in block.rows:
        for col, cell_data in enumerate(data_row, start=1):
            column = (
                block.columns[col - 1] if col - 1 < len(block.columns) else block.columns[-1]
            )
            if cell_data.display is not None:
                cell = ws.cell(row=row, column=col, value=cell_data.display)
                cell.alignment = WRAP
            else:
                cell = ws.cell(row=row, column=col, value=cell_data.value)
                number_format = excel_number_format(column.format, currency)
                if number_format:
                    cell.number_format = number_format
            if cell_data.emphasis:
                cell.font = BOLD
            fill = _hex(cell_data.color)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
        row += 1

    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(block.columns))}{row - 1}"
        if ws.auto_filter.ref is None
        else ws.auto_filter.ref
    )
    _widths(ws, [column.width for column in block.columns])

    if block.note:
        cell = ws.cell(row=row, column=1, value=block.note)
        cell.font = SUBTLE_FONT
        cell.alignment = WRAP
        row += 1
    return row + 1


def _write_keyvalues(ws: Worksheet, block: KeyValues, row: int) -> int:
    if block.caption:
        ws.cell(row=row, column=1, value=block.caption).font = CAPTION_FONT
        row += 1
    for item in block.items:
        ws.cell(row=row, column=1, value=item.label).font = BOLD
        ws.cell(row=row, column=2, value=item.value).alignment = WRAP
        if item.note:
            note = ws.cell(row=row, column=3, value=item.note)
            note.font = SUBTLE_FONT
            note.alignment = WRAP
        row += 1
    _widths(ws, [34, 30, 60])
    return row + 1


def _write_callout(ws: Worksheet, block: Callout, row: int) -> int:
    label = ws.cell(
        row=row, column=1, value=f"{TONE_LABEL.get(block.tone, 'NOTE')} — {block.title}"
    )
    label.font = BOLD
    label.fill = TONE_FILL.get(block.tone, TONE_FILL["info"])
    row += 1
    body = ws.cell(row=row, column=1, value=block.text)
    body.alignment = WRAP
    body.fill = TONE_FILL.get(block.tone, TONE_FILL["info"])
    _widths(ws, [34, 30, 60])
    return row + 2


def _write_matrix(ws: Worksheet, block: MatrixBlock, row: int) -> int:
    if block.caption:
        ws.cell(row=row, column=1, value=block.caption).font = CAPTION_FONT
        row += 1

    header = ws.cell(row=row, column=1, value="Probability \\ Impact")
    header.fill = HEADER_FILL
    header.font = HEADER_FONT
    for col, level in enumerate(block.impact_levels, start=2):
        cell = ws.cell(row=row, column=col, value=f"{level.level} · {level.label}")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for prob in block.probability_levels:
        head = ws.cell(row=row, column=1, value=f"{prob.level} · {prob.label}")
        head.font = BOLD
        tallest = 1
        for col, imp in enumerate(block.impact_levels, start=2):
            cell_data = block.cell(prob.level, imp.level)
            if cell_data is None:
                continue
            codes = list(cell_data.codes[:6])
            extra = len(cell_data.codes) - len(codes)
            lines = (
                [str(cell_data.count)] + codes + ([f"+{extra} more"] if extra else [])
                if cell_data.count
                else []
            )
            tallest = max(tallest, len(lines))
            cell = ws.cell(row=row, column=col, value="\n".join(lines))
            fill = _hex(cell_data.color)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = CENTER
        ws.row_dimensions[row].height = max(20, 13 * tallest)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Legend").font = BOLD
    row += 1
    for band in block.bands:
        swatch = ws.cell(row=row, column=1, value=band.name)
        fill = _hex(band.color)
        if fill:
            swatch.fill = PatternFill("solid", fgColor=fill)
        ws.cell(row=row, column=2, value=f"score {band.min_score} – {band.max_score}")
        row += 1

    if block.note:
        ws.cell(row=row, column=1, value=block.note).font = SUBTLE_FONT
        row += 1
    _widths(ws, [26] + [20] * len(block.impact_levels))
    return row + 1


def _write_section(ws: Worksheet, document: Document, section: Section) -> None:
    ws.cell(row=1, column=1, value=section.title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=document.subtitle).font = SUBTLE_FONT
    row = 4
    for block in section.blocks:
        if isinstance(block, Paragraph):
            cell = ws.cell(row=row, column=1, value=block.text)
            cell.alignment = WRAP
            _widths(ws, [34, 30, 60])
            row += 2
        elif isinstance(block, KeyValues):
            row = _write_keyvalues(ws, block, row)
        elif isinstance(block, Table):
            row = _write_table(ws, block, row, document.currency)
        elif isinstance(block, Callout):
            row = _write_callout(ws, block, row)
        elif isinstance(block, MatrixBlock):
            row = _write_matrix(ws, block, row)


def render_xlsx(document: Document) -> bytes:
    """The document as a workbook: a contents sheet, then one sheet per section."""
    wb = Workbook()
    contents = wb.active
    contents.title = "Contents"
    contents.cell(row=1, column=1, value=document.title).font = TITLE_FONT
    contents.cell(row=2, column=1, value=document.subtitle).font = SUBTLE_FONT
    contents.cell(
        row=3,
        column=1,
        value=(
            f"Prepared by {document.prepared_by} · {document.generated_on.isoformat()}"
            if document.prepared_by
            else document.generated_on.isoformat()
        ),
    ).font = SUBTLE_FONT

    used: set[str] = {"contents"}
    titles = [(_sheet_title(document, section, used), section) for section in
              document.sections]

    row = 5
    contents.cell(row=row, column=1, value="Sheet").font = HEADER_FONT
    contents.cell(row=row, column=1).fill = HEADER_FILL
    contents.cell(row=row, column=2, value="Section").font = HEADER_FONT
    contents.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    for title, section in titles:
        contents.cell(row=row, column=1, value=title)
        contents.cell(row=row, column=2, value=section.title)
        row += 1
    contents.cell(
        row=row + 1,
        column=1,
        value=format_value(
            "Figures are Monte Carlo estimates. The basis sheet states the seed, the "
            "iteration count and everything excluded from the run."
        ),
    ).font = SUBTLE_FONT
    contents.column_dimensions["A"].width = 34
    contents.column_dimensions["B"].width = 46

    for title, section in titles:
        _write_section(wb.create_sheet(title), document, section)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
