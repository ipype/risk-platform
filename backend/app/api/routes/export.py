"""Export of the risk register, mitigation actions, risk matrix, and RBS reference.

Three artifacts:
  ``GET /export/register.xlsx``  the full register workbook
  ``GET /export/matrix.xlsx``    the matrix for one lens and basis, plus placement and scale
  ``GET /export/matrix.svg``     the same matrix as a drop-into-a-report vector image

Placement is never computed here. It comes from ``app.services.matrix_export`` so the
workbook, the image and the screen always agree on which cell a risk sits in.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import Select, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.session import get_db
from app.models.custom_fields import get_custom_fields
from app.models.matrix import get_active_config
from app.models.mitigation import MitigationAction
from app.models.rbs import RbsCategory, RbsSubcategory
from app.models.risk import Risk
from app.services.matrix_export import (
    OVERALL,
    Grid,
    basis_label,
    build_grid,
    grid_to_svg,
    placement_for,
    valid_lens,
)

router = APIRouter(prefix="/export", tags=["export"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTLE_FONT = Font(color="6B7280", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CELL_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

RiskRow = tuple[Risk, RbsSubcategory, RbsCategory]


def _fmt(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value if value is not None else ""


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = f"A{row + 1}"


def _hex(color: str) -> str:
    return str(color).lstrip("#") or "FFFFFF"


# ------------------------------------------------------------------ register workbook


def _build_register_sheet(
    ws: Worksheet,
    risk_rows: list[RiskRow],
    areas: list[dict],
    custom_fields: list[dict],
    action_counts: dict[int, int],
    band_color: dict[str, str],
) -> None:
    area_codes = [a["code"] for a in areas]
    area_names = [a["name"] for a in areas]

    headers = (
        [
            "Risk Code",
            "Category",
            "Subcategory",
            "Title",
            "Description",
            "Causes",
            "Consequences",
            "Status",
            "Owner",
            "Last Review Date",
            "Current Probability",
            "Current Overall Impact",
            "Current Risk Level",
        ]
        + [f"Current {name}" for name in area_names]
        + [
            "Target Probability",
            "Target Overall Impact",
            "Target Risk Level",
        ]
        + [f"Target {name}" for name in area_names]
        + [f["label"] for f in custom_fields]
        + ["Mitigation Actions", "Comments", "Created At", "Updated At"]
    )
    _header_row(ws, headers)

    idx_current_level = headers.index("Current Risk Level") + 1
    idx_target_level = headers.index("Target Risk Level") + 1
    wrap_cols = [
        headers.index("Description") + 1,
        headers.index("Causes") + 1,
        headers.index("Consequences") + 1,
        headers.index("Comments") + 1,
    ]

    for risk, subcat, cat in risk_rows:
        cur_scores = risk.impact_scores or {}
        tgt_scores = risk.target_impact_scores or {}
        row = (
            [
                risk.risk_code,
                cat.name,
                f"{subcat.code} - {subcat.name}",
                risk.title,
                risk.description or "",
                risk.causes or "",
                risk.consequences or "",
                risk.status,
                risk.owner or "",
                _fmt(risk.last_review_date),
                risk.probability,
                risk.impact,
                risk.risk_level,
            ]
            + [cur_scores.get(code) for code in area_codes]
            + [
                risk.target_probability,
                risk.target_impact,
                risk.target_risk_level,
            ]
            + [tgt_scores.get(code) for code in area_codes]
            + [(risk.custom_fields or {}).get(f["key"]) for f in custom_fields]
            + [
                action_counts.get(risk.id, 0),
                risk.comments or "",
                _fmt(risk.created_at),
                _fmt(risk.updated_at),
            ]
        )
        ws.append(row)
        r = ws.max_row
        if risk.risk_level in band_color:
            ws.cell(row=r, column=idx_current_level).fill = PatternFill(
                "solid", fgColor=band_color[risk.risk_level]
            )
        if risk.target_risk_level in band_color:
            ws.cell(row=r, column=idx_target_level).fill = PatternFill(
                "solid", fgColor=band_color[risk.target_risk_level]
            )
        for col in wrap_cols:
            ws.cell(row=r, column=col).alignment = WRAP

    widths = [16, 20, 28, 30] + [40, 30, 30] + [12, 14, 16] + [16, 18, 16]
    widths += [14] * len(area_names) * 2
    widths += [16] * len(custom_fields)
    widths += [16, 40, 18, 18]
    _autosize(ws, widths[: len(headers)])


def _build_mitigations_sheet(
    ws: Worksheet, actions: list[tuple[MitigationAction, Risk]]
) -> None:
    headers = [
        "Risk Code",
        "Risk Title",
        "Action",
        "Owner",
        "Due Date",
        "Budget",
        "Completion %",
        "Effectiveness",
        "Status",
    ]
    _header_row(ws, headers)
    for action, risk in actions:
        ws.append(
            [
                risk.risk_code,
                risk.title,
                action.action,
                action.owner or "",
                _fmt(action.due_date),
                action.budget,
                action.completion_pct,
                action.effectiveness or "",
                action.status,
            ]
        )
        ws.cell(row=ws.max_row, column=3).alignment = WRAP
    _autosize(ws, [14, 30, 40, 16, 12, 12, 12, 14, 14])


def _build_rbs_sheet(ws: Worksheet, categories: list[RbsCategory]) -> None:
    headers = [
        "Category Code",
        "Category Name",
        "Subcategory Code",
        "Subcategory Name",
        "Description",
        "Full Prefix",
    ]
    _header_row(ws, headers)
    for cat in categories:
        for sub in cat.subcategories:
            ws.append(
                [
                    cat.code,
                    cat.name,
                    sub.code,
                    sub.name,
                    sub.description or "",
                    f"{cat.code}-{sub.code}",
                ]
            )
            ws.cell(row=ws.max_row, column=5).alignment = WRAP
    _autosize(ws, [14, 26, 16, 40, 50, 14])


# -------------------------------------------------------------------- matrix workbook


def _build_grid_sheet(
    ws: Worksheet,
    grid: Grid,
    bands: list[dict] | None = None,
    show_codes: bool = True,
    max_codes: int = 6,
) -> None:
    """The matrix itself: probability down the side, impact across the bottom of the read,
    laid out the same way as the screen so the two can be checked against each other."""
    ws.cell(row=1, column=1, value=f"Risk matrix — {grid.lens_label}").font = TITLE_FONT
    subtitle = basis_label(grid.basis)
    if grid.config_name:
        subtitle = f"{subtitle} · {grid.config_name}"
    ws.cell(row=2, column=1, value=subtitle).font = SUBTLE_FONT

    head = 4
    _header_row(
        ws,
        ["Probability \\ Impact"] + [f'{c["level"]} · {c["label"]}' for c in grid.columns],
        row=head,
    )

    for r_offset, prob in enumerate(grid.rows, start=1):
        row_idx = head + r_offset
        head_cell = ws.cell(
            row=row_idx, column=1, value=f'{prob["level"]} · {prob["label"]}'
        )
        head_cell.font = Font(bold=True)
        head_cell.alignment = Alignment(vertical="center")

        tallest = 1
        for c_offset, imp in enumerate(grid.columns, start=2):
            cell_data = grid.cell(prob["level"], imp["level"])
            codes = [p.code for p in cell_data.placements] if cell_data else []
            if not codes:
                value: str | int = ""
            elif show_codes:
                shown = codes[:max_codes]
                extra = len(codes) - len(shown)
                lines = [str(len(codes))] + shown + ([f"+{extra} more"] if extra else [])
                value = "\n".join(lines)
                tallest = max(tallest, len(lines))
            else:
                value = len(codes)

            cell = ws.cell(row=row_idx, column=c_offset, value=value)
            cell.fill = PatternFill("solid", fgColor=_hex(cell_data.color if cell_data else ""))
            cell.alignment = CELL_ALIGN

        ws.row_dimensions[row_idx].height = max(20, 13 * tallest)

    last_row = head + len(grid.rows)
    legend_row = last_row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = Font(bold=True)
    for i, band in enumerate(bands or [], start=1):
        swatch = ws.cell(row=legend_row + i, column=1, value=band["name"])
        swatch.fill = PatternFill("solid", fgColor=_hex(band["color"]))
        ws.cell(
            row=legend_row + i,
            column=2,
            value=f'score {band["min_score"]} – {band["max_score"]}',
        )

    footer_row = legend_row + len(bands or []) + 2
    summary = f"{len(grid.placed)} of {grid.total} risks placed"
    if grid.unplaced:
        summary += f" · {len(grid.unplaced)} not scored on this view (see Placement sheet)"
    ws.cell(row=footer_row, column=1, value=summary).font = SUBTLE_FONT
    ws.cell(
        row=footer_row + 1, column=1, value=f"Generated {date.today().isoformat()}"
    ).font = SUBTLE_FONT

    _autosize(ws, [26] + [20] * len(grid.columns))


def _build_placement_sheet(ws: Worksheet, grid: Grid, band_color: dict[str, str]) -> None:
    """Every risk and the cell it landed in — including the ones that landed nowhere."""
    headers = [
        "Risk Code",
        "Title",
        "Category",
        "Owner",
        "Status",
        "Probability",
        "Impact",
        "Score",
        "Band",
        "Placement",
    ]
    _header_row(ws, headers)
    for item in list(grid.placed) + list(grid.unplaced):
        if item.placed:
            placement = "Placed"
        elif item.off_scale:
            placement = "Off scale for the active config"
        else:
            placement = "Not scored on this view"
        ws.append(
            [
                item.code,
                item.title,
                item.category or "",
                item.owner or "",
                item.status,
                item.probability,
                item.impact,
                item.score,
                item.band or "",
                placement,
            ]
        )
        r = ws.max_row
        ws.cell(row=r, column=2).alignment = WRAP
        if item.band and item.band in band_color:
            ws.cell(row=r, column=9).fill = PatternFill(
                "solid", fgColor=band_color[item.band]
            )
    _autosize(ws, [16, 46, 20, 18, 12, 12, 10, 10, 14, 28])


def _build_scale_sheet(ws: Worksheet, config: dict) -> None:
    """The scoring scheme the matrix was drawn against — the appendix a reviewer asks for."""
    row = 1
    ws.cell(row=row, column=1, value="Probability levels").font = TITLE_FONT
    row += 1
    for level in sorted(config.get("probability_levels", []), key=lambda x: x["level"]):
        ws.cell(row=row, column=1, value=level["level"])
        ws.cell(row=row, column=2, value=level["label"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Impact levels").font = TITLE_FONT
    row += 1
    impact_levels = sorted(config.get("impact_levels", []), key=lambda x: x["level"])
    for level in impact_levels:
        ws.cell(row=row, column=1, value=level["level"])
        ws.cell(row=row, column=2, value=level["label"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Impact area descriptors").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Area").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for i, level in enumerate(impact_levels, start=2):
        cell = ws.cell(row=row, column=i, value=f'{level["level"]} · {level["label"]}')
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    for area in config.get("impact_areas", []):
        ws.cell(row=row, column=1, value=area["name"]).font = Font(bold=True)
        descriptors = area.get("descriptors", {}) or {}
        for i, level in enumerate(impact_levels, start=2):
            cell = ws.cell(row=row, column=i, value=descriptors.get(str(level["level"]), ""))
            cell.alignment = WRAP
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Bands").font = TITLE_FONT
    row += 1
    for band in config.get("bands", []):
        cell = ws.cell(row=row, column=1, value=band["name"])
        cell.fill = PatternFill("solid", fgColor=_hex(band["color"]))
        ws.cell(row=row, column=2, value=f'{band["min_score"]} – {band["max_score"]}')
        row += 1

    _autosize(ws, [22] + [26] * max(len(impact_levels), 1))


# ------------------------------------------------------------------------- data access


def _filtered(
    stmt: Select,
    category: str | None,
    status: str | None,
    owner: str | None,
) -> Select:
    if category:
        stmt = stmt.where(RbsCategory.code == category.strip().upper())
    if status:
        stmt = stmt.where(Risk.status == status)
    if owner:
        stmt = stmt.where(Risk.owner.ilike(f"%{owner.strip()}%"))
    return stmt


async def _load_risk_rows(
    db: AsyncSession,
    category: str | None = None,
    status: str | None = None,
    owner: str | None = None,
) -> list[RiskRow]:
    stmt = (
        select(Risk, RbsSubcategory, RbsCategory)
        .join(RbsSubcategory, RbsSubcategory.id == Risk.subcategory_id)
        .join(RbsCategory, RbsCategory.id == RbsSubcategory.category_id)
        .order_by(Risk.risk_code)
    )
    result = await db.execute(_filtered(stmt, category, status, owner))
    return list(result.all())


def _grid_from_rows(
    risk_rows: list[RiskRow], config: dict, lens: str, basis: str
) -> Grid:
    placements = [
        placement_for(risk, config, lens=lens, basis=basis, category=cat.name)
        for risk, _subcat, cat in risk_rows
    ]
    return build_grid(placements, config, lens=lens, basis=basis)


def _slug(lens: str) -> str:
    return "overall" if lens == OVERALL else lens.lower()


# ----------------------------------------------------------------------------- routes


@router.get("/register.xlsx")
async def export_register(db: AsyncSession = Depends(get_db)) -> StreamingResponse:
    risk_rows = await _load_risk_rows(db)

    actions = (
        (
            await db.execute(
                select(MitigationAction, Risk)
                .join(Risk, Risk.id == MitigationAction.risk_id)
                .order_by(Risk.risk_code, MitigationAction.sort_order)
            )
        )
        .all()
    )

    categories = (
        (
            await db.execute(
                select(RbsCategory)
                .options(selectinload(RbsCategory.subcategories))
                .order_by(RbsCategory.sort_order)
            )
        )
        .scalars()
        .all()
    )

    config = await get_active_config(db)
    custom_fields = (await get_custom_fields(db))["fields"]

    areas = config.get("impact_areas", [])
    bands = config.get("bands", [])
    band_color = {b["name"]: _hex(b["color"]) for b in bands}

    action_counts: dict[int, int] = {}
    for _action, risk in actions:
        action_counts[risk.id] = action_counts.get(risk.id, 0) + 1

    wb = Workbook()
    register_ws = wb.active
    register_ws.title = "Risk Register"
    _build_register_sheet(
        register_ws, risk_rows, areas, custom_fields, action_counts, band_color
    )

    _build_mitigations_sheet(wb.create_sheet("Mitigation Actions"), actions)
    _build_grid_sheet(
        wb.create_sheet("Risk Matrix"),
        _grid_from_rows(risk_rows, config, OVERALL, "current"),
        bands=bands,
    )
    _build_rbs_sheet(wb.create_sheet("RBS Reference"), categories)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"risk_register_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/matrix.xlsx")
async def export_matrix_xlsx(
    db: AsyncSession = Depends(get_db),
    lens: str = Query(default=OVERALL, description="Impact area code, or the overall lens"),
    basis: Literal["current", "target"] = Query(default="current"),
    category: str | None = Query(default=None, description="RBS category code"),
    status: str | None = Query(default=None),
    owner: str | None = Query(default=None),
    show_codes: bool = Query(default=True),
) -> StreamingResponse:
    config = await get_active_config(db)
    lens = valid_lens(lens, config)
    risk_rows = await _load_risk_rows(db, category, status, owner)
    grid = _grid_from_rows(risk_rows, config, lens, basis)
    band_color = {b["name"]: _hex(b["color"]) for b in config.get("bands", [])}

    wb = Workbook()
    grid_ws = wb.active
    grid_ws.title = "Risk Matrix"
    _build_grid_sheet(grid_ws, grid, bands=config.get("bands", []), show_codes=show_codes)
    _build_placement_sheet(wb.create_sheet("Placement"), grid, band_color)
    _build_scale_sheet(wb.create_sheet("Scale"), config)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"risk_matrix_{_slug(lens)}_{basis}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/matrix.svg")
async def export_matrix_svg(
    db: AsyncSession = Depends(get_db),
    lens: str = Query(default=OVERALL),
    basis: Literal["current", "target"] = Query(default="current"),
    category: str | None = Query(default=None),
    status: str | None = Query(default=None),
    owner: str | None = Query(default=None),
    show_codes: bool = Query(default=True),
    title: str = Query(default="Risk matrix", max_length=120),
) -> Response:
    config = await get_active_config(db)
    lens = valid_lens(lens, config)
    risk_rows = await _load_risk_rows(db, category, status, owner)
    grid = _grid_from_rows(risk_rows, config, lens, basis)

    svg = grid_to_svg(grid, project_title=title, show_codes=show_codes)
    filename = f"risk_matrix_{_slug(lens)}_{basis}_{date.today().isoformat()}.svg"
    return Response(
        content=svg,
        media_type="image/svg+xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
