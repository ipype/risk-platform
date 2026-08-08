"""Workbook extraction: one chunk per row, per sheet, carrying the header row.

**``data_only=True``, and this is the opposite of the convention everywhere else in the
repo.** The build-schedule workbook is opened with ``data_only=False`` so its formulas
survive a round trip; here a formula string is not evidence. ``=SUM(D4:D19)`` cited under a
risk about cost growth tells a reviewer nothing, and no retrieval will ever match it. What
is wanted is the number the estimator was looking at. The trade is real and worth naming:
a workbook saved by a tool that never computed its formulas has no cached values, so those
cells read as empty — which is why an empty sheet produces a warning rather than silence.

**The header row is guessed, and the guess is declared.** The first row with two or more
non-empty cells is taken as headers. Estimating workbooks routinely open with a title
block, a logo row, and two blank rows before anything tabular, and taking row 1 blindly
would label every column with a fragment of a title. Where no such row exists the rows are
still emitted, just without labels, and a warning says so.

**One chunk per row, never per sheet.** A sheet flattened into running text loses which
column a number sat under, which is generally the whole meaning of the number.
"""

from __future__ import annotations

import io
from typing import Any

from app.core.errors import DocumentHasNoText, DocumentUnreadable
from app.ingest.types import TABLE_ROW, Chunk, Extraction, row_text

SUFFIXES = (".xlsx", ".xlsm")

#: Sheets wider than this have their surplus columns dropped. A row carrying two hundred
#: labelled values is not a citable unit; it is a data export that belongs in a table, and
#: emitting it whole would fill the corpus with chunks no reviewer can read.
MAX_COLUMNS = 40

#: Guards against a sheet whose used range runs to the sheet limit because one cell far
#: down was once formatted. Reading it whole costs minutes and yields blank rows.
MAX_ROWS = 5000


def extract(data: bytes, *, filename: str = "") -> Extraction:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            io.BytesIO(data), data_only=True, read_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a wide family on bad zips
        raise DocumentUnreadable(filename or "workbook", str(exc)) from exc

    out = Extraction()
    try:
        for sheet in workbook.worksheets:
            _sheet(sheet, out)
    finally:
        workbook.close()

    if not out.chunks:
        raise DocumentHasNoText(
            filename or "workbook",
            "No readable values. Either the sheets are empty, or the file was written by "
            "a tool that never computed its formulas, so no cached results were stored.",
        )
    return out


def _sheet(sheet: Any, out: Extraction) -> None:
    rows: list[tuple[int, list[str]]] = []
    for offset, values in enumerate(
        sheet.iter_rows(values_only=True, max_row=MAX_ROWS), start=1
    ):
        cells = [_render(v) for v in values[:MAX_COLUMNS]]
        if any(cells):
            rows.append((offset, cells))

    if not rows:
        out.warnings.append(f"Sheet {sheet.title!r} held no values and was skipped.")
        return

    header_at, headers = _headers(rows)
    if headers is None:
        out.warnings.append(
            f"Sheet {sheet.title!r} has no row that looks like column headers, so its "
            "values are recorded without labels."
        )
        headers = []

    width = max(len(cells) for _, cells in rows)
    for number, cells in rows:
        if number == header_at:
            continue
        text = row_text(headers, cells)
        if not text:
            continue
        out.add(
            Chunk(
                kind=TABLE_ROW,
                text=text,
                locator={
                    "sheet": sheet.title,
                    "row": number,
                    "cells": f"A{number}:{_column(min(width, MAX_COLUMNS))}{number}",
                },
                section=sheet.title,
            )
        )


def _headers(rows: list[tuple[int, list[str]]]) -> tuple[int | None, list[str] | None]:
    for number, cells in rows:
        if sum(1 for c in cells if c.strip()) >= 2:
            return number, [c.strip() for c in cells]
    return None, None


def _render(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # 90.0 read back from a cell holding 90 is noise in every excerpt it appears in.
        return str(int(value))
    return str(value).strip()


def _column(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"
