"""Section builders, tested without a database.

That is the point of the split: a snapshot is constructed here by hand, so every assertion
about what a report says runs in milliseconds and none of them needs an engine, a schema
or a fixture. The end-to-end path — that the routes produce this snapshot from real rows —
is ``test_reports_api.py``.

Most of what is asserted here is not formatting. It is that the statements which stop a
reader misusing the numbers are actually printed: the additive-percentile gap, the gate
override, the excluded risks, and the approximation in the delay column.
"""

from __future__ import annotations

from datetime import date, datetime, timezone

import numpy as np
import pytest

from app.services.report import build_document, render_html, render_xlsx
from app.services.report.data import (
    ActionFacts,
    MatrixFacts,
    PlanFacts,
    ReductionFacts,
    ReportData,
    RiskFacts,
    RiskMoverFacts,
    RoiFacts,
    RunFacts,
    ScopeFacts,
    SeriesReductionFacts,
)
from app.services.report.model import (
    Callout,
    KeyValues,
    MatrixBand,
    MatrixCell,
    MatrixLevel,
    Table,
    format_value,
)
from app.services.report.sections import (
    SECTIONS,
    available_ids,
    build_sections,
    section_by_id,
)
from app.sim.engine import SimulationResult
from app.sim.joint import JointConfidence, JointFrontier, JointPoint
from app.sim.results import ContingencyView, DeterministicView, PercentilePoint, RunManifest
from app.sim.sensitivity import ActivityCriticality, RiskSensitivity

PERCENTILES = (10.0, 50.0, 80.0, 90.0)


def _series(values, label: str, units: str):
    from app.sim.results import summarise

    return summarise(
        np.asarray(values, dtype=np.float64),
        label=label,
        units=units,
        percentiles=PERCENTILES,
        s_curve_points=11,
        histogram_bins=5,
    )


def _result(*, with_schedule: bool = True, additive_gap: bool = True) -> SimulationResult:
    rng = np.random.default_rng(11)
    risk_cost = rng.uniform(100_000, 900_000, size=500)
    delay = rng.uniform(-3, 40, size=500)
    burn = delay * 10_000
    total = 5_000_000 + risk_cost + (burn if with_schedule else 0.0)

    contingency = ContingencyView(
        base_cost=5_000_000.0,
        mean_total_cost=float(total.mean()),
        contingency=tuple(
            PercentilePoint(p=p, value=float(np.percentile(total, p)) - 5_000_000.0)
            for p in PERCENTILES
        ),
        additive_error_at_p80=41_000.0 if additive_gap else None,
        additive_p80_total=float(np.percentile(total, 80)) + 41_000.0
        if additive_gap
        else None,
        integrated_p80_total=float(np.percentile(total, 80)) if additive_gap else None,
        cost_variance_share=0.62,
        schedule_variance_share=0.38,
    )

    return SimulationResult(
        manifest=RunManifest(
            engine_version="1.4.0",
            seed=4242,
            iterations=500,
            sampling="lhs",
            centered_lhs=True,
            chunk_size=250,
            inputs_sha256="a" * 64,
        ),
        deterministic=DeterministicView(
            base_cost=5_000_000.0,
            activities=140 if with_schedule else 0,
            relationships=210 if with_schedule else 0,
            inserted_activities=3 if with_schedule else 0,
            baseline_finish_day=420.0 if with_schedule else None,
            critical_activities=22 if with_schedule else 0,
        ),
        contingency=contingency,
        risk_cost=_series(risk_cost, "Risk cost", "currency"),
        total_cost=_series(total, "Total cost", "currency"),
        delay_days=_series(delay, "Schedule delay", "days") if with_schedule else None,
        finish_day=_series(delay + 420, "Project finish", "days") if with_schedule else None,
        schedule_driven_cost=_series(burn, "Burn-rate cost", "currency")
        if with_schedule
        else None,
        risk_sensitivity=(
            RiskSensitivity(
                risk_id=1,
                code="TEC-DES-0001",
                title="Design growth",
                cost_variance_share=0.31,
                schedule_variance_share=0.12 if with_schedule else None,
                combined_variance_share=0.43,
                spearman_total_cost=0.55,
                mean_contribution=310_000.0,
                p80_contribution=520_000.0,
                realised_frequency=0.61,
            ),
            RiskSensitivity(
                risk_id=2,
                code="EXT-PER-0002",
                title="Permit delay",
                cost_variance_share=0.09,
                schedule_variance_share=0.14 if with_schedule else None,
                combined_variance_share=0.23,
                spearman_total_cost=0.31,
                mean_contribution=90_000.0,
                p80_contribution=180_000.0,
                realised_frequency=0.4,
            ),
        ),
        joint=(
            JointConfidence(
                iterations=500,
                frontiers=(
                    JointFrontier(
                        target=80.0,
                        points=(
                            JointPoint(
                                delay_days=31.0,
                                finish_day=451.0,
                                total_cost=6_100_000.0,
                                delay_p=88.0,
                                cost_p=88.0,
                            ),
                        ),
                        balanced=JointPoint(
                            delay_days=31.0,
                            finish_day=451.0,
                            total_cost=6_100_000.0,
                            delay_p=88.0,
                            cost_p=88.0,
                        ),
                    ),
                ),
                marginal_pair_target=80.0,
                marginal_cost=5_900_000.0,
                marginal_delay_days=28.0,
                joint_at_marginal_pair=0.68,
                cost_delay_correlation=0.74,
                burn_rate_coupled=True,
            )
            if with_schedule
            else None
        ),
        schedule_variance_share=0.38 if with_schedule else 0.0,
        activity_criticality=(
            (
                ActivityCriticality(
                    activity_id="A1000",
                    code="A1000",
                    name="Detailed design",
                    criticality_index=0.94,
                    mean_total_float_days=1.2,
                    duration_sensitivity=0.71,
                    cruciality=0.67,
                    duration_sd_days=6.4,
                    schedule_sensitivity_index=0.55,
                ),
                ActivityCriticality(
                    activity_id="A2000",
                    code="A2000",
                    name="Permit approval",
                    criticality_index=0.41,
                    mean_total_float_days=18.0,
                    duration_sensitivity=None,
                    cruciality=0.0,
                    duration_sd_days=0.0,
                    schedule_sensitivity_index=0.0,
                    is_inserted=True,
                ),
            )
            if with_schedule
            else ()
        ),
        warnings=("One risk was sampled at fewer than 30 realised occurrences.",),
    )


def _run(**overrides) -> RunFacts:
    base = dict(
        id=7,
        name="August baseline",
        status="succeeded",
        scenario="pre_mitigation",
        created_by="Sam",
        created_at=datetime(2026, 8, 4, 9, 30, tzinfo=timezone.utc),
        finished_at=datetime(2026, 8, 4, 9, 32, tzinfo=timezone.utc),
        duration_ms=124_000,
        iterations=500,
        seed=4242,
        sampling="lhs",
        engine_version="1.4.0",
        chunk_size=250,
        inputs_sha256="a" * 64,
        base_cost=5_000_000.0,
        burn_rate_per_day=10_000.0,
        risk_count=12,
        mapped_risk_count=5,
        activity_count=140,
        schedule_version_id=3,
        gate_passed=True,
        gate_override=False,
        excluded=(
            {
                "risk_id": 9,
                "risk_code": "COM-SUP-0009",
                "title": "Supplier insolvency",
                "reason": "cost_max: maximum is below the most likely value",
            },
        ),
        assembly_notes=("Activities span 2 calendars; durations converted to elapsed days.",),
        result=_result(),
    )
    base.update(overrides)
    return RunFacts(**base)


def _risks() -> tuple[RiskFacts, ...]:
    return (
        RiskFacts(
            id=1,
            code="TEC-DES-0001",
            title="Design growth",
            category="Technical",
            subcategory="DES — Design",
            status="open",
            owner="A. Engineer",
            probability=4,
            impact=5,
            score=20,
            band="High",
            band_color="#dc2626",
            quantified=True,
        ),
        RiskFacts(
            id=2,
            code="EXT-PER-0002",
            title="Permit delay",
            category="External",
            subcategory="PER — Permitting",
            status="open",
            probability=3,
            impact=3,
            score=9,
            band="Medium",
            band_color="#f59e0b",
            quantified=True,
        ),
        RiskFacts(
            id=3,
            code="COM-SUP-0009",
            title="Supplier insolvency",
            category="Commercial",
            subcategory="SUP — Supply chain",
            status="open",
        ),
    )


def _matrix() -> MatrixFacts:
    return MatrixFacts(
        lens_label="Overall (worst area)",
        basis_label="Pre-mitigation",
        config_name="Default 5x5",
        probability_levels=(
            MatrixLevel(level=4, label="Likely"),
            MatrixLevel(level=3, label="Possible"),
        ),
        impact_levels=(
            MatrixLevel(level=3, label="Moderate"),
            MatrixLevel(level=5, label="Severe"),
        ),
        cells=(
            MatrixCell(
                probability=4,
                impact=5,
                score=20,
                band="High",
                color="#dc2626",
                count=1,
                codes=("TEC-DES-0001",),
            ),
            MatrixCell(
                probability=3,
                impact=3,
                score=9,
                band="Medium",
                color="#f59e0b",
                count=1,
                codes=("EXT-PER-0002",),
            ),
        ),
        bands=(
            MatrixBand(name="Medium", color="#f59e0b", min_score=6, max_score=14),
            MatrixBand(name="High", color="#dc2626", min_score=15, max_score=25),
        ),
        placed=2,
        unplaced=1,
    )


def _data(**overrides) -> ReportData:
    base = dict(
        title="Quantitative risk analysis report",
        subtitle="Portfolio › Metro Extension",
        prepared_by="Sam",
        currency="$",
        generated_on=date(2026, 8, 5),
        scope=ScopeFacts(
            id=2, kind="project", name="Metro Extension", path=("Portfolio", "Metro Extension")
        ),
        risks=_risks(),
        matrix=_matrix(),
        run=_run(),
        actions=(
            ActionFacts(
                risk_code="TEC-DES-0001",
                risk_title="Design growth",
                action="Freeze the design basis at 60%",
                owner="A. Engineer",
                budget=120_000.0,
                sched_days=0.0,
                completion_pct=40,
                status="in_progress",
            ),
        ),
    )
    base.update(overrides)
    return ReportData(**base)


def _sections(data: ReportData) -> dict[str, object]:
    return {section.id: section for section in build_sections(data)}


def _text(section) -> str:
    """Every string a section carries, flattened. Cheap way to assert a statement lands."""
    parts: list[str] = [section.title]
    for block in section.blocks:
        parts.append(getattr(block, "text", "") or "")
        parts.append(getattr(block, "title", "") or "")
        parts.append(getattr(block, "caption", "") or "")
        parts.append(getattr(block, "note", "") or "")
        if isinstance(block, KeyValues):
            for item in block.items:
                parts += [item.label, item.value, item.note or ""]
        if isinstance(block, Table):
            parts += [column.label for column in block.columns]
            parts.append(block.empty_text)
            for row in block.rows:
                parts += [str(cell.display or cell.value or "") for cell in row]
    return "\n".join(parts)


class TestAvailability:
    def test_a_full_run_makes_every_section_available(self):
        data = _data(plan=PlanFacts(id=1, name="Package A", status="materialized"))
        ids = available_ids(data)
        assert set(ids) == {spec.id for spec in SECTIONS}

    def test_registry_order_is_the_document_order(self):
        data = _data()
        built = [section.id for section in build_sections(data)]
        assert built == [
            spec.id for spec in SECTIONS if spec.unavailable(data) is None
        ]

    def test_requesting_sections_filters_but_does_not_reorder(self):
        data = _data()
        built = [s.id for s in build_sections(data, ["cost", "cover", "basis"])]
        assert built == ["cover", "basis", "cost"]

    def test_unknown_section_id_is_ignored_rather_than_raising(self):
        assert [s.id for s in build_sections(_data(), ["cover", "nonsense"])] == ["cover"]

    def test_a_cost_only_run_says_why_the_schedule_sections_are_missing(self):
        data = _data(run=_run(result=_result(with_schedule=False), activity_count=0))
        assert "cost only" in (section_by_id("schedule").unavailable(data) or "")
        assert section_by_id("criticality").unavailable(data) is not None
        assert "schedule" not in available_ids(data)
        # the cost half is untouched by the absence
        assert "cost" in available_ids(data)

    def test_no_run_leaves_a_register_only_report(self):
        data = _data(run=None)
        assert set(available_ids(data)) == {
            "cover",
            "basis",
            "method",
            "register",
            "matrix",
            "actions",
        }
        assert "No simulation run was selected." == section_by_id("cost").unavailable(data)

    def test_an_empty_register_still_reports_its_basis(self):
        data = _data(risks=(), matrix=None, actions=(), run=None)
        ids = available_ids(data)
        assert "register" not in ids and "matrix" not in ids
        assert {"cover", "basis", "method"}.issubset(set(ids))

    def test_a_run_whose_result_would_not_parse_degrades_to_the_basis(self):
        data = _data(
            run=_run(result=None, result_error="The stored result could not be read back.")
        )
        assert "cost" not in available_ids(data)
        assert "unreadable" in _text(_sections(data)["basis"]).lower()


class TestBasisStatesWhatItRests_On:
    def test_the_reproducibility_record_is_printed(self):
        text = _text(_sections(_data())["basis"])
        for expected in ("Seed", "4242", "Iterations", "1.4.0", "a" * 64, "LHS"):
            assert expected in text

    def test_excluded_risks_are_named_not_counted(self):
        text = _text(_sections(_data())["basis"])
        assert "COM-SUP-0009" in text
        assert "maximum is below the most likely value" in text

    def test_an_empty_exclusion_list_says_so_rather_than_showing_nothing(self):
        data = _data(run=_run(excluded=()))
        assert "No risk was excluded" in _text(_sections(data)["basis"])

    def test_a_gate_override_is_a_warning_at_the_top(self):
        data = _data(
            run=_run(
                gate_passed=False,
                gate_override=True,
                gate_override_reason="Baseline reissue lands next week.",
            )
        )
        basis = _sections(data)["basis"]
        first = basis.blocks[0]
        assert isinstance(first, Callout)
        assert first.tone == "warning"
        assert "Baseline reissue lands next week." in first.text

    def test_a_scope_note_from_gathering_is_carried_into_the_report(self):
        data = _data(notes=("A scope was requested that is not the run's own.",))
        assert "not the run's own" in _text(_sections(data)["basis"])


class TestTheArithmeticIsStatedOnTheFace:
    def test_method_leads_with_the_additive_percentile_rule(self):
        method = _sections(_data())["method"]
        first = method.blocks[0]
        assert isinstance(first, Callout)
        assert first.tone == "method"
        assert "never added" in first.title.lower()

    def test_the_additive_gap_is_printed_next_to_the_integrated_figure(self):
        text = _text(_sections(_data())["cost"])
        assert "41,000" in text
        assert "never the number to use" in text

    def test_no_additive_gap_means_no_claim_about_one(self):
        data = _data(run=_run(result=_result(additive_gap=False)))
        assert "never the number to use" not in _text(_sections(data)["cost"])

    def test_the_apportionment_approximation_is_declared_where_it_is_used(self):
        text = _text(_sections(_data())["drivers"])
        assert "apportioned" in text.lower()
        assert "exact" in text.lower()

    def test_burn_rate_cost_is_not_presented_as_delay_times_rate(self):
        text = _text(_sections(_data())["schedule"])
        assert "not the P80 delay multiplied by the burn rate" in text

    def test_delay_is_declared_unclamped(self):
        assert "unclamped" in _text(_sections(_data())["schedule"])

    def test_the_joint_section_prices_the_marginal_pair(self):
        joint = _sections(_data())["joint"]
        first = joint.blocks[0]
        assert isinstance(first, Callout)
        # quoting a P80 cost beside a P80 date implies 68%, not 80%
        assert "68.0%" in first.text
        table = next(b for b in joint.blocks if isinstance(b, Table))
        assert [row[0].display for row in table.rows] == ["Joint P80"]


class TestContent:
    def test_the_cost_headline_carries_contingency_and_its_share_of_base(self):
        cost = _sections(_data())["cost"]
        headline = next(b for b in cost.blocks if isinstance(b, KeyValues))
        labels = {item.label for item in headline.items}
        assert {"Base cost", "Contingency at P80", "Contingency as share of base"} <= labels

    def test_drivers_are_ranked_by_combined_share_not_by_cost_share(self):
        drivers = _sections(_data())["drivers"]
        table = next(b for b in drivers.blocks if isinstance(b, Table))
        assert [row[1].display for row in table.rows] == ["TEC-DES-0001", "EXT-PER-0002"]

    def test_a_missing_duration_sensitivity_prints_as_missing_not_as_zero(self):
        criticality = _sections(_data())["criticality"]
        table = next(b for b in criticality.blocks if isinstance(b, Table))
        permit = next(row for row in table.rows if row[0].display == "A2000")
        assert permit[5].value is None
        assert format_value(permit[5].value, "ratio") == "—"
        # ... while a measured zero stays a zero
        assert permit[4].value == 0.0

    def test_the_register_separates_scored_from_quantified(self):
        register = _sections(_data())["register"]
        summary = next(b for b in register.blocks if isinstance(b, KeyValues))
        values = {item.label: item.value for item in summary.items}
        assert values["Risks in register"] == "3"
        assert values["Qualitatively scored"] == "2"
        assert values["Carrying a quantitative estimate"] == "2"

    def test_the_matrix_carries_band_names_beside_the_colours(self):
        matrix = _sections(_data())["matrix"]
        block = matrix.blocks[0]
        assert {band.name for band in block.bands} == {"Medium", "High"}
        assert block.cell(4, 5).codes == ("TEC-DES-0001",)
        assert "not scored on this view" in (block.note or "")


class TestRenderers:
    def test_html_is_one_self_contained_file(self):
        html = render_html(build_document(_data()))
        assert html.startswith("<!doctype html>")
        assert "<style>" in html and "src=" not in html
        assert "@page" in html  # printable
        assert 'lang="en"' in html

    def test_html_escapes_register_content(self):
        risky = RiskFacts(
            id=4,
            code="TEC-DES-0004",
            title="<script>alert('x')</script>",
            category="Technical",
            status="open",
            probability=5,
            impact=5,
            score=25,
        )
        html = render_html(build_document(_data(risks=_risks() + (risky,))))
        assert "<script>alert" not in html
        assert "&lt;script&gt;" in html

    def test_html_carries_every_built_section_and_a_contents_entry(self):
        data = _data()
        document = build_document(data)
        html = render_html(document)
        for section in document.sections:
            assert f'id="{section.id}"' in html
            assert f'href="#{section.id}"' in html

    def test_xlsx_opens_and_keeps_money_as_numbers(self):
        from io import BytesIO

        from openpyxl import load_workbook

        document = build_document(_data())
        wb = load_workbook(BytesIO(render_xlsx(document)))
        assert wb.sheetnames[0] == "Contents"
        assert "Cost contingency" in wb.sheetnames

        sheet = wb["Cost contingency"]
        numeric = [
            cell
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, (int, float)) and cell.value > 1_000_000
        ]
        assert numeric, "the cost table wrote no numbers"
        assert any('"$"#,##0' in (cell.number_format or "") for cell in numeric)

    def test_xlsx_sheet_titles_stay_inside_excels_limit(self):
        document = build_document(_data())
        wb_names = render_xlsx(document)
        from io import BytesIO

        from openpyxl import load_workbook

        names = load_workbook(BytesIO(wb_names)).sheetnames
        assert len(names) == len(set(names))
        assert all(len(name) <= 31 for name in names)

    def test_an_empty_document_is_refused_upstream_not_rendered_blank(self):
        data = _data(risks=(), matrix=None, actions=(), run=None)
        document = build_document(data, ["cost", "criticality"])
        assert document.sections == ()


class TestFormatting:
    @pytest.mark.parametrize(
        ("value", "fmt", "expected"),
        [
            (None, "currency", "—"),
            (1_234_567.0, "currency", "$1,234,567"),
            (12.26, "days", "12.3 d"),
            (0.615, "pct", "61.5%"),
            (0.0, "pct", "0.0%"),
            (-3.5, "days", "-3.5 d"),
            (1.234, "ratio", "1.23"),
        ],
    )
    def test_values_format_the_same_wherever_they_are_printed(self, value, fmt, expected):
        assert format_value(value, fmt, "$") == expected

    def test_a_blank_currency_prints_plain_numbers(self):
        assert format_value(1000.0, "currency", "") == "1,000"


def _reduction(before: float, after: float) -> ReductionFacts:
    cut = before - after
    return ReductionFacts(
        before=before, after=after, reduction=cut, reduction_pct=cut / before
    )


def _roi(**overrides) -> RoiFacts:
    base = dict(
        id=3,
        name="Package A vs baseline",
        plan_id=1,
        percentile=80.0,
        seed_shared=True,
        before_run_id=7,
        after_run_id=8,
        status="ready",
        contingency=SeriesReductionFacts(
            label="Contingency",
            units="currency",
            mean=_reduction(640_000.0, 470_000.0),
            at_percentile=_reduction(910_000.0, 640_000.0),
            standard_error=18_000.0,
            within_noise=False,
        ),
        delay_days=SeriesReductionFacts(
            label="Schedule delay",
            units="days",
            mean=_reduction(21.0, 14.0),
            at_percentile=_reduction(34.0, 22.0),
        ),
        plan_budget=120_000.0,
        plan_sched_days=0.0,
        benefit_cost_ratio=2.25,
        net_at_percentile=150_000.0,
        retired_count=2,
        risk_movers=(
            RiskMoverFacts(
                code="TEC-DES-0001",
                title="Design growth",
                movement="reduced",
                contribution_before=310_000.0,
                contribution_after=120_000.0,
                contribution_reduction=190_000.0,
                rank_before=1,
                rank_after=3,
            ),
            RiskMoverFacts(
                code="COM-SUP-0009",
                title="Supplier insolvency",
                movement="retired",
                contribution_before=80_000.0,
                contribution_after=0.0,
                contribution_reduction=80_000.0,
                rank_before=4,
                rank_after=None,
            ),
        ),
        basis=("Both runs share a seed, so the delta is not reading sampling noise.",),
        warnings=(),
    )
    base.update(overrides)
    return RoiFacts(**base)


class TestMitigationSection:
    """The one section whose whole job is a comparison, so the one most able to mislead."""

    def test_a_plan_alone_reports_the_package_without_claiming_an_effect(self):
        data = _data(
            plan=PlanFacts(
                id=1,
                name="Package A",
                status="materialized",
                action_count=6,
                costed_count=5,
                unpriced_count=1,
                total_budget=120_000.0,
            )
        )
        text = _text(_sections(data)["mitigation"])
        assert "Package A" in text
        assert "Benefit / cost" not in text
        assert "cost-side twin of dropping a risk from a run" in text

    def test_effectiveness_is_declared_as_a_re_simulation_delta(self):
        section = _sections(_data(roi=_roi()))["mitigation"]
        callout = next(b for b in section.blocks if isinstance(b, Callout))
        assert callout.tone == "method"
        assert "difference between two full runs" in callout.text

    def test_the_package_cost_sits_beside_the_contingency_not_inside_it(self):
        section = _sections(_data(roi=_roi()))["mitigation"]
        headline = next(
            b for b in section.blocks if isinstance(b, KeyValues) and b.caption == "At P80"
        )
        values = {item.label: item.value for item in headline.items}
        assert values["Contingency before"] == "$910,000"
        assert values["Contingency after"] == "$640,000"
        assert values["Reduction"] == "$270,000"
        assert values["Package budget"] == "$120,000"
        assert values["Net at this percentile"] == "$150,000"

    def test_a_reduction_inside_the_error_bar_says_so_against_the_number(self):
        roi = _roi(
            contingency=SeriesReductionFacts(
                label="Contingency",
                units="currency",
                mean=_reduction(640_000.0, 638_000.0),
                at_percentile=_reduction(910_000.0, 908_000.0),
                standard_error=18_000.0,
                within_noise=True,
            )
        )
        section = _sections(_data(roi=roi))["mitigation"]
        headline = next(
            b for b in section.blocks if isinstance(b, KeyValues) and b.caption == "At P80"
        )
        reduction = next(item for item in headline.items if item.label == "Reduction")
        assert reduction.note is not None
        assert "noise" in reduction.note

    def test_an_invalidated_pair_prints_the_issues_instead_of_the_numbers(self):
        roi = _roi(issues=("The residual register was re-materialised after this pair ran.",))
        text = _text(_sections(_data(roi=roi))["mitigation"])
        assert "no longer valid" in text
        assert "re-materialised" in text
        assert "$910,000" not in text
        assert "Benefit / cost" not in text

    def test_the_movers_table_names_what_was_retired(self):
        section = _sections(_data(roi=_roi()))["mitigation"]
        table = next(
            b
            for b in section.blocks
            if isinstance(b, Table) and b.caption == "Where the reduction came from"
        )
        assert [row[0].display for row in table.rows] == ["TEC-DES-0001", "COM-SUP-0009"]
        assert [row[2].display for row in table.rows] == ["Reduced", "Retired"]

    def test_a_populated_comparison_renders_in_both_formats(self):
        from io import BytesIO

        from openpyxl import load_workbook

        document = build_document(
            _data(roi=_roi(), plan=PlanFacts(id=1, name="Package A", status="materialized"))
        )
        html = render_html(document)
        assert "Mitigation and its effect" in html
        assert "$270,000" in html

        wb = load_workbook(BytesIO(render_xlsx(document)))
        assert "Mitigation and its effect" in wb.sheetnames
