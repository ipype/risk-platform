"""End-to-end tests for the report routes.

Self-contained in the style of ``test_roi_api.py``: its own app, its own SQLite schema,
``simulation_eager`` so a real run exists to report on. What is under test here is the
join — that ``gather`` reads the right rows, that naming a run pins the scope, and that
all three renderings come out of the same build. The content of a section is asserted
against a hand-built snapshot in ``test_report_sections.py``, which is far cheaper.

Two projects are seeded and most of the interesting assertions are about the boundary
between them. A report that quietly printed another project's register underneath this
project's contingency would look entirely plausible, which is exactly why it is worth a
test rather than a comment.
"""

from __future__ import annotations

from io import BytesIO

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.api.errors import register_exception_handlers
from app.api.routes import reports as report_routes
from app.api.routes import simulations as simulation_routes
from app.core.config import settings
from app.db.base_class import Base
from app.db.session import get_db
from app.models.mapping import RiskActivityMapping
from app.models.mitigation import MitigationAction
from app.models.quant import RiskQuantEstimate
from app.models.rbs import RbsCategory, RbsSubcategory
from app.models.risk import Risk
from app.models.roi import MitigationRoi  # noqa: F401  (registers the table)
from app.models.scope import ScopeNode
from app.models.schedule import (
    DcmaRun,
    ScheduleActivity,
    ScheduleFile,
    ScheduleRelationship,
    ScheduleVersion,
)
from app.models.simulation import SimulationRun  # noqa: F401  (registers the table)

PORTFOLIO = 1
PROJECT_A = 2
PROJECT_B = 3

FAST = {"iterations": 400, "seed": 7}

COST = {
    "bound_interpretation": "absolute",
    "cost_dist": "pert",
    "cost_min": 200_000.0,
    "cost_ml": 400_000.0,
    "cost_max": 900_000.0,
    "confidence": "high",
}


def _activity(source_id: str, code: str, days: float, *, kind: str = "task"):
    return ScheduleActivity(
        version_id=1,
        source_id=source_id,
        code=code,
        name=f"Activity {code}",
        calendar_source_id="CAL-1",
        wbs_source_id="W1",
        type=kind,
        status="not_started",
        duration_calendar_id="CAL-1",
        original_duration_days=days,
        remaining_duration_days=days,
        total_float_days=0.0,
    )


async def _seed(session) -> None:
    session.add(
        ScopeNode(id=PORTFOLIO, kind="portfolio", name="Capital portfolio", created_by="test")
    )
    session.add(
        ScopeNode(
            id=PROJECT_A,
            kind="project",
            parent_id=PORTFOLIO,
            name="Metro Extension",
            is_default=True,
            created_by="test",
        )
    )
    session.add(
        ScopeNode(
            id=PROJECT_B,
            kind="project",
            parent_id=PORTFOLIO,
            name="Depot Upgrade",
            created_by="test",
        )
    )
    session.add(RbsCategory(id=1, code="TEC", name="Technical"))
    session.add(RbsSubcategory(id=1, category_id=1, code="DES", name="Design"))

    rows = [
        (1, PROJECT_A, "TEC-DES-0001", "Scope growth", 4, 5),
        (2, PROJECT_A, "TEC-DES-0002", "Ground conditions", 3, 3),
        (3, PROJECT_A, "TEC-DES-0003", "Half-elicited", None, None),
        (4, PROJECT_B, "TEC-DES-0004", "Someone else's risk", 5, 5),
    ]
    for risk_id, scope, code, title, probability, impact in rows:
        session.add(
            Risk(
                id=risk_id,
                scope_id=scope,
                subcategory_id=1,
                seq=risk_id,
                risk_code=code,
                title=title,
                probability=probability,
                impact=impact,
                owner="A. Analyst",
            )
        )

    session.add(
        RiskQuantEstimate(risk_id=1, scenario="pre_mitigation", p_occurrence=0.5, **COST)
    )
    session.add(
        RiskQuantEstimate(
            risk_id=2,
            scenario="pre_mitigation",
            p_occurrence=0.4,
            sched_dist="pert",
            sched_min=5.0,
            sched_ml=10.0,
            sched_max=30.0,
            sched_day_basis="working",
            **COST,
        )
    )
    # A shape with no numbers behind it. It must appear in the basis section as excluded,
    # never be quietly absent from a contingency computed over the rest.
    session.add(
        RiskQuantEstimate(
            risk_id=3,
            scenario="pre_mitigation",
            p_occurrence=0.3,
            bound_interpretation="absolute",
            cost_dist="pert",
        )
    )
    session.add(
        RiskQuantEstimate(risk_id=4, scenario="pre_mitigation", p_occurrence=0.9, **COST)
    )

    session.add(
        MitigationAction(
            risk_id=1,
            action="Freeze the design basis at 60%",
            owner="A. Engineer",
            budget=120_000.0,
            completion_pct=40,
            status="in_progress",
        )
    )
    session.add(
        MitigationAction(
            risk_id=4,
            action="Another project's action",
            owner="Nobody here",
            status="open",
        )
    )

    session.add(
        ScheduleFile(
            id=1,
            scope_id=PROJECT_A,
            filename="test.xer",
            suffix=".xer",
            content=b"x",
            content_sha256="a" * 64,
            size_bytes=1,
        )
    )
    session.add(
        ScheduleVersion(
            id=1,
            file_id=1,
            source_project_id="P1",
            project_name="Metro Extension",
            source_format="xer",
            parser_version="1.0",
            activity_count=4,
            relationship_count=3,
        )
    )
    for row in (
        _activity("A1", "A1000", 5.0),
        _activity("A2", "A1010", 10.0),
        _activity("A3", "A1020", 7.0),
        _activity("A4", "A1030", 0.0, kind="milestone_finish"),
    ):
        session.add(row)
    for i, (pred, succ) in enumerate((("A1", "A2"), ("A2", "A3"), ("A3", "A4")), start=1):
        session.add(
            ScheduleRelationship(
                version_id=1,
                source_id=f"R{i}",
                predecessor_source_id=pred,
                successor_source_id=succ,
                type="FS",
                lag_days=0.0,
            )
        )
    session.add(
        RiskActivityMapping(
            risk_id=2,
            version_id=1,
            mapping_type="duration_driver",
            activity_source_id="A2",
            status="accepted",
            origin="manual",
        )
    )
    session.add(
        DcmaRun(
            version_id=1,
            gate_passed=True,
            passed_count=14,
            failed_count=0,
            not_assessed_count=0,
        )
    )
    await session.commit()


@pytest_asyncio.fixture
async def client(monkeypatch):
    monkeypatch.setattr(settings, "simulation_eager", True)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    maker = async_sessionmaker(engine, expire_on_commit=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    async with maker() as session:
        await _seed(session)

    app = FastAPI()
    register_exception_handlers(app)
    app.include_router(simulation_routes.router)
    app.include_router(report_routes.router)

    async def _override():
        async with maker() as session:
            yield session

    app.dependency_overrides[get_db] = _override
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    await engine.dispose()


async def _run(client, **overrides) -> dict:
    payload = {"name": "August baseline", "base_cost": 5_000_000.0, **FAST, **overrides}
    response = await client.post("/simulations", json=payload)
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "succeeded", body.get("error")
    return body


async def _scheduled_run(client) -> dict:
    return await _run(
        client, schedule_version_id=1, burn_rate_per_day=25_000.0
    )


# --------------------------------------------------------------------------------------
# the section manifest
# --------------------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_sections_answers_for_the_selected_run_not_in_general(client):
    cost_only = await _run(client)
    body = (await client.get("/reports/sections", params={"run_id": cost_only["id"]})).json()
    by_id = {section["id"]: section for section in body["sections"]}

    assert by_id["cost"]["available"] is True
    assert by_id["schedule"]["available"] is False
    assert "cost only" in by_id["schedule"]["reason"]
    assert by_id["criticality"]["available"] is False
    assert body["run_id"] == cost_only["id"]
    assert body["scope_id"] == PROJECT_A

    scheduled = await _scheduled_run(client)
    body = (await client.get("/reports/sections", params={"run_id": scheduled["id"]})).json()
    by_id = {section["id"]: section for section in body["sections"]}
    assert by_id["schedule"]["available"] is True
    assert by_id["criticality"]["available"] is True


@pytest.mark.asyncio
async def test_sections_without_a_run_offers_the_register_half(client):
    body = (await client.get("/reports/sections")).json()
    by_id = {section["id"]: section for section in body["sections"]}

    assert by_id["register"]["available"] is True
    assert by_id["matrix"]["available"] is True
    assert by_id["cost"]["available"] is False
    assert by_id["cost"]["reason"] == "No simulation run was selected."
    assert by_id["mitigation"]["available"] is False


@pytest.mark.asyncio
async def test_a_run_that_does_not_exist_is_a_404_not_an_empty_report(client):
    response = await client.get("/reports/sections", params={"run_id": 9999})
    assert response.status_code == 404
    assert "9999" in response.text


# --------------------------------------------------------------------------------------
# scope
# --------------------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_the_register_in_a_report_is_the_runs_own_project(client):
    run = await _run(client)
    body = (await client.get("/reports/report.json", params={"run_id": run["id"]})).json()
    payload = str(body)

    assert "TEC-DES-0001" in payload
    assert "TEC-DES-0004" not in payload, "another project's register reached this report"
    assert "Another project's action" not in payload


@pytest.mark.asyncio
async def test_naming_a_wider_scope_than_the_run_is_recorded_not_obeyed(client):
    run = await _run(client)
    body = (
        await client.get(
            "/reports/sections", params={"run_id": run["id"], "scope_id": PORTFOLIO}
        )
    ).json()

    assert body["scope_id"] == PROJECT_A
    assert body["notes"], "the ignored scope must be said out loud"
    assert "run's own" in body["notes"][0]


@pytest.mark.asyncio
async def test_without_a_run_a_scope_reports_everything_under_it(client):
    body = (
        await client.get("/reports/report.json", params={"scope_id": PORTFOLIO})
    ).json()
    payload = str(body)
    assert "TEC-DES-0001" in payload and "TEC-DES-0004" in payload

    body = (await client.get("/reports/report.json", params={"scope_id": PROJECT_B})).json()
    payload = str(body)
    assert "TEC-DES-0004" in payload and "TEC-DES-0001" not in payload


# --------------------------------------------------------------------------------------
# the document
# --------------------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_the_document_carries_the_basis_of_its_own_numbers(client):
    run = await _scheduled_run(client)
    body = (
        await client.get(
            "/reports/report.json",
            params={"run_id": run["id"], "currency": "$", "prepared_by": "Sam"},
        )
    ).json()

    basis = next(section for section in body["sections"] if section["id"] == "basis")
    payload = str(basis)
    assert run["inputs_sha256"] in payload
    assert str(run["seed"]) in payload
    assert run["engine_version"] in payload
    # the risk with a shape and no numbers is named, not silently absent
    assert "TEC-DES-0003" in payload


@pytest.mark.asyncio
async def test_sections_are_filtered_but_not_reordered_by_the_request(client):
    run = await _run(client)
    body = (
        await client.get(
            "/reports/report.json",
            params=[("run_id", run["id"]), ("section", "cost"), ("section", "cover")],
        )
    ).json()
    assert [section["id"] for section in body["sections"]] == ["cover", "cost"]


@pytest.mark.asyncio
async def test_asking_only_for_sections_with_nothing_to_say_is_refused(client):
    response = await client.get(
        "/reports/report.json", params=[("section", "criticality")]
    )
    assert response.status_code == 422
    assert "/reports/sections" in response.text


@pytest.mark.asyncio
async def test_the_cost_section_reports_the_integrated_contingency(client):
    run = await _scheduled_run(client)
    body = (
        await client.get(
            "/reports/report.json", params={"run_id": run["id"], "currency": "$"}
        )
    ).json()

    cost = next(section for section in body["sections"] if section["id"] == "cost")
    headline = next(block for block in cost["blocks"] if block["kind"] == "keyvalues")
    labels = {item["label"] for item in headline["items"]}
    assert {"Base cost", "Contingency at P80"} <= labels

    table = next(block for block in cost["blocks"] if block["kind"] == "table")
    p80 = next(row for row in table["rows"] if row[0]["display"] == "P80")
    contingency = p80[2]["value"]
    total = p80[1]["value"]
    # the contingency is the integrated total less the base, not a sum of parts
    assert total == pytest.approx(5_000_000.0 + contingency, rel=1e-9)


# --------------------------------------------------------------------------------------
# renderings
# --------------------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_html_is_one_printable_self_contained_file(client):
    run = await _scheduled_run(client)
    response = await client.get(
        "/reports/report.html", params={"run_id": run["id"], "currency": "$"}
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/html")
    assert "content-disposition" not in response.headers
    html = response.text
    assert html.startswith("<!doctype html>")
    assert "@page" in html
    assert "TEC-DES-0001" in html
    assert "TEC-DES-0004" not in html


@pytest.mark.asyncio
async def test_html_can_be_asked_for_as_a_download(client):
    run = await _run(client)
    response = await client.get(
        "/reports/report.html", params={"run_id": run["id"], "download": True}
    )
    disposition = response.headers["content-disposition"]
    assert "attachment" in disposition
    assert "metro-extension" in disposition


@pytest.mark.asyncio
async def test_the_workbook_opens_with_a_sheet_per_section(client):
    run = await _scheduled_run(client)
    response = await client.get(
        "/reports/report.xlsx", params={"run_id": run["id"], "currency": "$"}
    )

    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames[0] == "Contents"
    for expected in ("Basis of the analysis", "Cost contingency", "Schedule outcome"):
        assert expected in wb.sheetnames
    assert all(len(name) <= 31 for name in wb.sheetnames)


@pytest.mark.asyncio
async def test_every_rendering_comes_from_the_same_build(client):
    run = await _scheduled_run(client)
    params = {"run_id": run["id"], "currency": "$"}

    document = (await client.get("/reports/report.json", params=params)).json()
    html = (await client.get("/reports/report.html", params=params)).text
    wb = load_workbook(
        BytesIO((await client.get("/reports/report.xlsx", params=params)).content)
    )

    titles = [section["title"] for section in document["sections"]]
    for title in titles:
        assert title in html
        assert title[:31] in wb.sheetnames


@pytest.mark.asyncio
async def test_a_register_only_report_renders_without_a_run(client):
    response = await client.get("/reports/report.html", params={"scope_id": PROJECT_A})
    assert response.status_code == 200
    assert "No simulation attached" in response.text
    assert "Cost contingency" not in response.text
